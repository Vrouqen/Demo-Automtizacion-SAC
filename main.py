import os
from io import BytesIO
from typing import Dict, Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pymongo import MongoClient
from pymongo.collection import Collection


app = FastAPI(title="Credenciales API")


def parse_credentials_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo Excel está vacío")

    credentials: Dict[str, Any] = {}

    start_row = 0
    first_row = rows[0]
    if first_row and len(first_row) >= 2:
        first_key = str(first_row[0]).strip().lower() if first_row[0] is not None else ""
        first_value = str(first_row[1]).strip().lower() if first_row[1] is not None else ""
        if first_key in {"key", "clave"} and first_value in {"value", "valor"}:
            start_row = 1

    for row in rows[start_row:]:
        if not row or len(row) < 2:
            continue
        key = row[0]
        value = row[1]
        if key is None:
            continue
        key_text = str(key).strip()
        if not key_text:
            continue
        credentials[key_text] = value

    if not credentials:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron credenciales válidas (columnas: key/value)",
        )

    return credentials


def get_collection() -> Collection:
    mongodb_uri = os.getenv("MONGODB_URI")
    if not mongodb_uri:
        raise HTTPException(status_code=500, detail="Falta configurar MONGODB_URI")

    db_name = os.getenv("MONGODB_DB", "sac")
    collection_name = os.getenv("MONGODB_COLLECTION", "credentials")

    client = MongoClient(mongodb_uri)
    return client[db_name][collection_name]


@app.post("/credentials/{credential_id}")
async def upload_credentials(credential_id: str, file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xlsx)")

    file_bytes = await file.read()
    credentials_dict = parse_credentials_excel(file_bytes)

    collection = get_collection()
    result = collection.update_one(
        {"_id": credential_id},
        {"$set": {"credentials": credentials_dict}},
        upsert=True,
    )

    return {
        "id": credential_id,
        "credentials_count": len(credentials_dict),
        "upserted": result.upserted_id is not None,
    }
